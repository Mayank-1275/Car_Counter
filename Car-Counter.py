import numpy as np
from ultralytics import YOLO
import cv2
import cvzone
import math
from sort import *
import schedule
import time
from openpyxl import load_workbook
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def Write_Data(text):

    # Load the existing Excel workbook
    workbook = load_workbook('data.xlsx')

    # Select the active sheet
    sheet = workbook.active

    # Get the current date and time
    current_datetime = datetime.datetime.now()
    next_row = sheet.max_row + 1

    # Append the current date, time, and text to the Excel sheet
    sheet.cell(row=next_row, column=1).value = current_datetime.date()
    sheet.cell(row=next_row, column=2).value = current_datetime.time()
    sheet.cell(row=next_row, column=3).value = text

    # Save the workbook
    workbook.save('data.xlsx')

    print(str(text)+"Data written to Excel sheet.")

def Email(data):

    # getting current date and time
    current_datetime = datetime.datetime.now()

    # Email information
    sender_email = "#######"
    receiver_email = "#######"
    subject = "Number Of Cars"
    message = ("Number of cars passed so far is "+ str(data)) +" at " + str(current_datetime)

    # Create a MIME multipart message
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg["Subject"] = subject

    # Attach the message to the MIME message
    msg.attach(MIMEText(message, "plain"))

    # SMTP server configuration
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    username = "#######"
    password = "#######"

    # Create an SMTP session
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()  # Enable secure connection
            server.login(username, password)  # Login to your email account
            server.send_message(msg)  # Send the email
        print("Email sent successfully!")

    # if any problem occures
    except smtplib.SMTPException as e:
        print("Error: Unable to send email. ", e)

Main program strats from here

cap = cv2.VideoCapture("cars.mp4")  # For Video
# cap=cv2.VideoCapture(0) if you want access your webcam


model = YOLO("yolov8l.pt")

classNames = ["person", "bicycle", "car", "motorbike", "aeroplane", "bus", "train", "truck", "boat",
              "traffic light", "fire hydrant", "stop sign", "parking meter", "bench", "bird", "cat",
              "dog", "horse", "sheep", "cow", "elephant", "bear", "zebra", "giraffe", "backpack", "umbrella",
              "handbag", "tie", "suitcase", "frisbee", "skis", "snowboard", "sports ball", "kite", "baseball bat",
              "baseball glove", "skateboard", "surfboard", "tennis racket", "bottle", "wine glass", "cup",
              "fork", "knife", "spoon", "bowl", "banana", "apple", "sandwich", "orange", "broccoli",
              "carrot", "hot dog", "pizza", "donut", "cake", "chair", "sofa", "pottedplant", "bed",
              "diningtable", "toilet", "tvmonitor", "laptop", "mouse", "remote", "keyboard", "cell phone",
              "microwave", "oven", "toaster", "sink", "refrigerator", "book", "clock", "vase", "scissors",
              "teddy bear", "hair drier", "toothbrush"
              ]

mask = cv2.imread("mask.png")

# Tracking
tracker = Sort(max_age=20, min_hits=3, iou_threshold=0.3)

limits = [400, 297, 673, 297]
totalCount = []

while True:
    success, img = cap.read()
    imgregion= cv2.bitwise_and(img, mask)

    imgGraphics = cv2.imread("graphics.png", cv2.IMREAD_UNCHANGED)
    img = cvzone.overlayPNG(img, imgGraphics, (0, 0))
    results = model(imgregion, stream=True)

    detections = np.empty((0, 5))

    for r in results:
        boxes = r.boxes
        for box in boxes:
            # Bounding Box
            x1, y1, x2, y2 = box.xyxy[0]
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            w, h = x2 - x1, y2 - y1

            # Confidence
            conf = math.ceil((box.conf[0] * 100)) / 100

            # Class Name
            cls = int(box.cls[0])
            currentClass = classNames[cls]

            if currentClass == "car" or currentClass == "truck" or currentClass == "bus" \
                    or currentClass == "motorbike" and conf > 0.3:
                currentArray = np.array([x1, y1, x2, y2, conf])
                detections = np.vstack((detections, currentArray))

    resultsTracker = tracker.update(detections)

    cv2.line(img, (limits[0], limits[1]), (limits[2], limits[3]), (0, 0, 255), 5)
    for result in resultsTracker:
        x1, y1, x2, y2, id = result
        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
        # print(result)
        w, h = x2 - x1, y2 - y1
        cvzone.cornerRect(img, (x1, y1, w, h), l=9, rt=2, colorR=(255, 0, 255))
        cvzone.putTextRect(img, f' {int(id)}', (max(0, x1), max(35, y1)),scale=2, thickness=3, offset=10)

        cx, cy = x1 + w // 2, y1 + h // 2
        cv2.circle(img, (cx, cy), 5, (255, 0, 255), cv2.FILLED)

        if limits[0] < cx < limits[2] and limits[1] - 15 < cy < limits[1] + 15:
            if totalCount.count(id) == 0:
                totalCount.append(id)
                cv2.line(img, (limits[0], limits[1]), (limits[2], limits[3]), (0, 255, 0), 5)

    cv2.putText(img,str(len(totalCount)),(255,100),cv2.FONT_HERSHEY_PLAIN,5,(50,50,255),8)

    # If you want to automate these two function at a perticuler time
    # schedule.every().day.at("12:00").do(Write_Data(len(totalCount)))
    # schedule.every().day.at("12:00").do(Email(len(totalCount)))

    # Feed the current number of vehicles in XML sheet
    if cv2.waitKey(1) & 0xFF == ord('s'):
        Write_Data(len(totalCount))

    # Send the current number of vehicles in XML sheet
    # if cv2.waitKey(1) & 0xFF == ord('e'):
    #     Email(len(totalCount))

    cv2.imshow("Image", img)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Terminate The Program
cv2.destroyAllWindows()
